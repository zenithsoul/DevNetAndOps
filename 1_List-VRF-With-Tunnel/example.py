import json
import requests
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, PatternFill, Border, Side
import urllib3

# IP of the Device IOS-XE
IP_Device = "127.0.0.1"
# Authentication credentials
username = "example"
password = "example"
# Output Excel
output_file = "example.xlsx"


# Disable insecure request warnings
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

url = "https://" + IP_Device + "/restconf/data/Cisco-IOS-XE-vrf-oper:vrf-oper-data"  # Replace with the actual HTTPS URL
headers = {
    "Accept": "application/yang-data+json"
}

# Disable SSL certificate verification (not recommended for production use)
response = requests.get(url, auth=(username, password), headers=headers, verify=False)
response.raise_for_status()  # Raise an exception if the request was unsuccessful

# Define two PatternFill instances for the colors you want
yellow_fill = PatternFill(start_color="F6FA70",
                   end_color="F6FA70",
                   fill_type="solid")
skyblue_fill = PatternFill(start_color="00DFA2",
                   end_color="00DFA2",
                   fill_type="solid")

try:
    data = response.json()
    vrf_entries = data["Cisco-IOS-XE-vrf-oper:vrf-oper-data"]["vrf-entry"]

    # Create a DataFrame to store the collected data
    data_df = pd.DataFrame(columns=["VRF name", "Interfaces", "IP", "Mask", "Source Tunnel", "Destination Tunnel"])

    for entry in vrf_entries:
        vrf_name = entry["vrf-name"]
        interfaces = entry.get("interface", [])

        # Check if interfaces list is not empty
        if interfaces:
            for interface in interfaces:
                num_tunnel = interface.replace("Tunnel", "").replace("tunnel", "")
                url_tunnel = f"https://" + IP_Device + "/restconf/data/Cisco-IOS-XE-native:native/interface/Tunnel={num_tunnel}"
                res_tunnel = requests.get(url_tunnel, auth=(username, password), headers=headers, verify=False)
                res_tunnel.raise_for_status()
                data_tunnel = res_tunnel.json()

                ip_tunnel = data_tunnel["Cisco-IOS-XE-native:Tunnel"]["ip"]["address"]["primary"]["address"]
                mask_tunnel = data_tunnel["Cisco-IOS-XE-native:Tunnel"]["ip"]["address"]["primary"].get("mask", "")
                src_tunnel = data_tunnel["Cisco-IOS-XE-native:Tunnel"]["Cisco-IOS-XE-tunnel:tunnel"]["source"]
                des_tunnel = data_tunnel["Cisco-IOS-XE-native:Tunnel"]["Cisco-IOS-XE-tunnel:tunnel"]["destination-config"]["ipv4"]

                data_df = data_df.append(
                    {"VRF name": vrf_name, "Interfaces": interface, "IP": ip_tunnel, "Mask": mask_tunnel, "Source Tunnel": src_tunnel, "Destination Tunnel": des_tunnel},
                    ignore_index=True
                )
        else:
            data_df = data_df.append(
                {"VRF name": vrf_name, "Interfaces": None, "IP": None, "Mask": None, "Source Tunnel": None, "Destination Tunnel": None},
                ignore_index=True
            )

except json.JSONDecodeError as e:
    print("Failed to decode JSON response:", e)
    print("Response Content:", response.content)

workbook = Workbook()
sheet = workbook.active

data_rows = list(dataframe_to_rows(data_df, index=False, header=True))

# Initialize a dict to track the index of each unique VRF name
vrf_indices = {}

for data_row in data_rows:
    sheet.append(data_row)
    vrf_name = data_row[0]
    
    if vrf_name not in vrf_indices:
        vrf_indices[vrf_name] = len(vrf_indices)
    
    vrf_index = vrf_indices[vrf_name]

    fill = yellow_fill if vrf_index % 2 == 0 else skyblue_fill
    for cell in sheet[sheet.max_row]:
        cell.fill = fill

previous_vrf_name = ""
start_row = 2
end_row = 2  
for row in sheet.iter_rows(min_row=2, min_col=1, max_col=1):
    current_vrf_name = row[0].value
    if current_vrf_name == previous_vrf_name:
        end_row = row[0].row
    else:
        if start_row != end_row:
            sheet.merge_cells(f"A{start_row}:A{end_row}")
        start_row = row[0].row
        end_row = row[0].row
    previous_vrf_name = current_vrf_name

for row in sheet.iter_rows(min_row=2, min_col=1, max_col=1):
    for cell in row:
        cell.alignment = Alignment(vertical="center")

# Define a border style
border = Border(left=Side(border_style='thin', color='000000'),
                right=Side(border_style='thin', color='000000'),
                top=Side(border_style='thin', color='000000'),
                bottom=Side(border_style='thin', color='000000'))

# Apply border to all cells
for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
    for cell in row:
        cell.border = border


workbook.save(output_file)

print("Data exported to", output_file)
