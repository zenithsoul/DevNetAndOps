import pandas as pd
import requests
import urllib3
from openpyxl.styles import PatternFill, Border, Side, Alignment
from openpyxl import Workbook

# Disable insecure request warnings
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# Replace with the actual URL of the JSON data
IP_Device = "127.0.0.1"

# Authentication credentials
username = "example"
password = "example"

# Output Excel
output_file  = "example.xlsx"

headers = {
    "Accept": "application/yang-data+json"
}

url = "https://" + IP_Device + "/restconf/data/Cisco-IOS-XE-native:native/interface/Tunnel"  

# Disable SSL certificate verification (not recommended for production use)
response = requests.get(url, auth=(username, password), headers=headers, verify=False)
response.raise_for_status()  # Raise an exception if the request was unsuccessful

if response.status_code == 200:
    json_data = response.json()

    # Extract the data
    names = [item['name'] for item in json_data['Cisco-IOS-XE-native:Tunnel']]
    ip_addresses = [
        {
            'IP': item['ip']['address']['primary']['address'],
            'Mask': item['ip']['address']['primary']['mask']
        }
        for item in json_data['Cisco-IOS-XE-native:Tunnel']
    ]
    sources = [item['Cisco-IOS-XE-tunnel:tunnel']['source'] for item in json_data['Cisco-IOS-XE-native:Tunnel']]
    destinations = [item['Cisco-IOS-XE-tunnel:tunnel']['destination-config']['ipv4'] for item in json_data['Cisco-IOS-XE-native:Tunnel']]
    forwarding_words = [item['ip']['vrf']['forwarding']['word'] if 'ip' in item and 'vrf' in item['ip'] and 'forwarding' in item['ip']['vrf'] else "-" for item in json_data['Cisco-IOS-XE-native:Tunnel']]

    # Create a DataFrame from the extracted data
    df = pd.DataFrame({
        'Tunnel Number': names,
        'IP Address': [ip['IP'] for ip in ip_addresses],
        'Mask': [ip['Mask'] for ip in ip_addresses],
        'Source': sources,
        'Destination': destinations,
        'VRF': forwarding_words
    })

    # Create a new workbook
    workbook = Workbook()
    sheet = workbook.active

    # Set the column names in the first row
    column_names = list(df.columns)
    for col_num, column_name in enumerate(column_names, start=1):
        sheet.cell(row=1, column=col_num).value = column_name

    # Set the data from the DataFrame
    for row_num, row_data in enumerate(df.values, start=2):
        for col_num, cell_value in enumerate(row_data, start=1):
            sheet.cell(row=row_num, column=col_num).value = cell_value

    # Apply conditional formatting to highlight rows with "-" in forwarding_words
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        if row[-1].value == "-":
            for cell in row:
                cell.fill = red_fill

    # Add black border to all columns and rows
    max_col = sheet.max_column
    max_row = sheet.max_row

    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    for row in sheet.iter_rows(min_row=1, max_row=max_row):
        for cell in row:
            cell.border = thin_border

    # Set the alignment to middle and center
    for row in sheet.iter_rows(min_row=1, max_row=max_row):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    
    workbook.save(output_file)

    print("Data has been written to " + output_file)

else:
    print("Error: Failed to retrieve data from the URL")
